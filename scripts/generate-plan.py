from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

GOLD = 'F59E0B'
PURPLE = '7C3AED'
GREEN = '22C55E'
BLUE = '3B82F6'
RED = 'EF4444'
DARK = '1A1A2E'
WHITE = 'FFFFFF'
GRAY = '9CA3AF'
LIGHT_GOLD = 'FEF3C7'
LIGHT_PURPLE = 'EDE9FE'
LIGHT_GREEN = 'DCFCE7'
LIGHT_BLUE = 'DBEAFE'

header_font = Font(bold=True, color=WHITE, size=11)
title_font = Font(bold=True, color=GOLD, size=16)
subtitle_font = Font(bold=True, color=PURPLE, size=12)
normal_font = Font(size=10)
thin_border = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill_color):
    fill = PatternFill('solid', fgColor=fill_color)
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, cols, fill_color=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = left_wrap
        cell.border = thin_border
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)

# ═══════════════════════════════════
# HOJA 1: IMPLEMENTADO
# ═══════════════════════════════════
ws1 = wb.active
ws1.title = 'Implementado'
ws1.sheet_properties.tabColor = GREEN
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 55
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 14

ws1.merge_cells('A1:E1')
ws1['A1'] = 'COPA OMEGA STAR - Features Implementadas'
ws1['A1'].font = title_font
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:E2')
ws1['A2'] = 'Beyblade X Tournament App - Bladers Santa Fe | copa-omega-rho.vercel.app'
ws1['A2'].font = Font(italic=True, color=GRAY, size=10)
ws1['A2'].alignment = center

r = 4
for c, h in enumerate(['Feature', 'Categoria', 'Descripcion', 'Fecha', 'Estado'], 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 5, DARK)

core = [
    ('Auth (registro/login)', 'Core', 'Registro y login con Supabase Auth, sin confirmacion de email', 'Pre-sprint', 'Prod'),
    ('Dashboard', 'Core', 'Stats del jugador, podio, racha de victorias, partidas recientes, accesos rapidos', 'Pre-sprint', 'Prod'),
    ('Perfiles de jugador', 'Core', 'Avatar con crop, beys (blade/ratchet/bit), tagline, emoji badge, color de acento', 'Pre-sprint', 'Prod'),
    ('Perfiles publicos', 'Core', 'Stats, head-to-head, historial de matches, badges, podium cards', 'Pre-sprint', 'Prod'),
    ('Sistema de estrellas', 'Core', 'Arrancan con 25, apuestan 1-5 por batalla, eliminados al llegar a 0', 'Pre-sprint', 'Prod'),
    ('Torneos (3 formatos)', 'Core', 'Eliminacion directa, Round Robin, Suizo. Registro QR, bracket, jueces, podio', 'Pre-sprint', 'Prod'),
    ('Retos (Challenges)', 'Core', 'Desafios 1v1 con apuesta de estrellas, 48hs para responder', 'Pre-sprint', 'Prod'),
    ('Feed social', 'Core', 'Actividad en tiempo real con reacciones y comentarios en batallas', 'Pre-sprint', 'Prod'),
    ('Combos', 'Core', 'Compartir combinaciones blade/ratchet/bit con votacion up/down', 'Pre-sprint', 'Prod'),
    ('Predicciones', 'Core', 'Predecir ganadores de matches, leaderboard de accuracy', 'Pre-sprint', 'Prod'),
    ('Encuestas', 'Core', 'Crear polls, votar, ver resultados', 'Pre-sprint', 'Prod'),
    ('Chat + Bot IA', 'Core', 'Chat global en tiempo real con BeyBot (Groq AI)', 'Pre-sprint', 'Prod'),
    ('Tienda', 'Core', 'Productos con imagenes, carrito, pago efectivo/transferencia', 'Pre-sprint', 'Prod'),
    ('Galeria', 'Core', 'Fotos y videos por torneo', 'Pre-sprint', 'Prod'),
    ('Modo espectador', 'Core', 'Overlay fullscreen para proyector/TV con VS screen y bracket', 'Pre-sprint', 'Prod'),
    ('Streaks + titulos', 'Core', 'Racha de login diario, titulos dinamicos por ultimas 10 partidas', 'Pre-sprint', 'Prod'),
    ('Admin panel', 'Core', 'Gestion de matches, torneos, jugadores, productos, pedidos, carousel', 'Pre-sprint', 'Prod'),
]

r = 5
for feat in core:
    for c, val in enumerate(feat, 1):
        ws1.cell(r, c, val)
    style_row(ws1, r, 5, LIGHT_GREEN if r % 2 == 0 else None)
    r += 1

ws1.merge_cells(f'A{r}:E{r}')
ws1.cell(r, 1, 'SPRINT 16-17 MARZO 2026')
ws1.cell(r, 1).font = subtitle_font
ws1.cell(r, 1).alignment = center
ws1.cell(r, 1).fill = PatternFill('solid', fgColor=LIGHT_PURPLE)
ws1.row_dimensions[r].height = 28
r += 1

sprint = [
    ('Push Notifications', 'Engagement', 'Notificaciones push: reto recibido/aceptado, resultado partida, podio torneo', '16/03', 'Prod'),
    ('Cron Cleanup', 'Backend', 'Expira challenges +48hs, desactiva polls vencidas. Diario 6am UTC', '16/03', 'Prod'),
    ('Rivalidades', 'Social', 'Detecta rivales (3+ matches). W/L, dominancia%, estrellas intercambiadas', '16/03', 'Prod'),
    ('12 Badges nuevos (23 total)', 'Gamificacion', 'Fenix, Gladiador, Oraculo, David, Centurion, Combo Master, Social King...', '16/03', 'Prod'),
    ('Star Transaction Log', 'Backend', 'Historial automatico de cada cambio de estrellas con trigger en DB', '16/03', 'Prod'),
    ('Auto-Seeding', 'Torneos', 'Siembra inteligente: estrellas*2 + wins*3 - losses + puntos torneo', '16/03', 'Prod'),
    ('Live Bracket (Realtime)', 'Torneos', 'Bracket se actualiza solo sin refresh. Indicador "En vivo"', '16/03', 'Prod'),
    ('Share Cards', 'Social', 'Imagen PNG 600x400 con stats para WhatsApp/Instagram', '16/03', 'Prod'),
    ('Check-in Pre-Torneo', 'Torneos', 'Ventana de check-in. Solo confirmados entran al bracket', '16/03', 'Prod'),
    ('Llave 3er Puesto', 'Torneos', 'Match de bronce entre perdedores de semi. Winner=3ro, Loser=4to', '16/03', 'Prod'),
    ('Editar torneos finalizados', 'Admin', 'Editar nombre, fecha, resultados, re-avanzar byes', '16/03', 'Prod'),
    ('Recalcular puntos', 'Admin', 'Boton para borrar y recalcular puntos/badges del torneo', '16/03', 'Prod'),
    ('Podio manual', 'Admin', 'Asignar 1ro/2do/3ro a cualquier jugador via API', '16/03', 'Prod'),
    ('Reordenar torneos', 'Admin', 'Flechas arriba/abajo para cambiar orden en la lista', '16/03', 'Prod'),
    ('Asignar jugadores al bracket', 'Admin', 'Boton "+" en slots vacios para asignar jugador por alias', '17/03', 'Prod'),
    ('Podio en Ranking Torneos', 'UI', 'Top 3 cards en pestana Torneos (igual que Estrellas)', '16/03', 'Prod'),
    ('Onboarding Checklist', 'UX', '5 pasos: foto, bey, ranking, prediccion, reto. Barra progreso', '17/03', 'Prod'),
    ('Temporadas (backend)', 'Sistema', 'Crear/iniciar/completar seasons. Snapshot + reset estrellas. Banner dashboard', '17/03', 'Prod'),
    ('Fix Combo Voting', 'Bugfix', 'RLS bloqueaba UPDATE de contadores - resuelto con admin client', '17/03', 'Fix'),
    ('Fix Challenges Stuck', 'Bugfix', 'Challenges en "accepted" - ahora se marcan completed', '16/03', 'Fix'),
    ('Fix Champion Overlay', 'Bugfix', 'Campeon flotaba encima de FINALIZADO - ahora barra debajo', '16/03', 'Fix'),
    ('86 E2E Tests', 'Testing', '71 flujos + 15 edge cases torneos. Playwright Chromium+Mobile', '16/03', 'Repo'),
]

for feat in sprint:
    for c, val in enumerate(feat, 1):
        ws1.cell(r, c, val)
    style_row(ws1, r, 5, LIGHT_GOLD if r % 2 == 0 else None)
    r += 1

# ═══════════════════════════════════
# HOJA 2: PENDIENTE P2
# ═══════════════════════════════════
ws2 = wb.create_sheet('Pendiente P2')
ws2.sheet_properties.tabColor = PURPLE
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 12

ws2.merge_cells('A1:E1')
ws2['A1'] = 'COPA OMEGA STAR - Proximo Sprint (P2)'
ws2['A1'].font = title_font
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 35

r = 3
for c, h in enumerate(['Feature', 'Descripcion', 'Impacto Esperado', 'Esfuerzo', 'Prioridad'], 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 5, PURPLE)

p2 = [
    ('Misiones Diarias', '3 misiones rotativas: "Gana 1 partida", "Vota en encuesta", "Comparte combo". Checklist en dashboard. 1 mision semanal.', 'DAU +40%. Razon para abrir la app todos los dias', '3-4 dias', 'ALTA'),
    ('Niveles / XP', 'XP que solo sube: ganar=+20, perder=+5, torneo=+10. Niveles tematicos: Principiante a Omega. Cosmeticos desbloqueables.', 'Retencion. Progresion sin perder estrellas', '3-4 dias', 'ALTA'),
    ('Historial Detallado', 'Tipo de victoria (burst/spin/over/xtreme). Combo usado por partida. Stats por tipo de victoria.', 'Profundidad competitiva. Datos para meta-game', '2-3 dias', 'MEDIA'),
    ('Ranking con Filtros', 'Filtrar por semana/mes/hoy. Sub-rankings: mejor win rate, mejor racha. Seccion "Movers".', 'Visibilidad para jugadores que mejoran', '1-2 dias', 'MEDIA'),
    ('Resumen Semanal', 'Modal los lunes: batallas, estrellas, posicion. Boton "Compartir" genera imagen.', 'Engagement pasivo. Contenido social', '1 dia', 'MEDIA'),
    ('Revancha Instantanea', 'Boton post-batalla para retar al mismo rival con mismas estrellas.', 'Reduce friccion. Mas partidas', '0.5 dias', 'BAJA'),
    ('Eventos Especiales', 'Admin crea eventos con countdown: "Weekend Warrior", "Combo Challenge".', 'Picos de actividad programables', '2 dias', 'MEDIA'),
    ('Landing Page Redesign', 'Rediseno completo con estetica Beyblade X. Hero, features, podio, CTA. Mobile-first.', 'Primera impresion. Conversion registro', '3-5 dias', 'ALTA'),
]

r = 4
for feat in p2:
    for c, val in enumerate(feat, 1):
        ws2.cell(r, c, val)
    style_row(ws2, r, 5, LIGHT_PURPLE if r % 2 == 0 else None)
    prio = ws2.cell(r, 5)
    if feat[4] == 'ALTA':
        prio.font = Font(bold=True, color=RED, size=10)
    elif feat[4] == 'MEDIA':
        prio.font = Font(bold=True, color=GOLD, size=10)
    else:
        prio.font = Font(bold=True, color=BLUE, size=10)
    r += 1

# ═══════════════════════════════════
# HOJA 3: PENDIENTE P3
# ═══════════════════════════════════
ws3 = wb.create_sheet('Pendiente P3')
ws3.sheet_properties.tabColor = BLUE
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 45

ws3.merge_cells('A1:D1')
ws3['A1'] = 'COPA OMEGA STAR - Vision a Futuro (P3)'
ws3['A1'].font = title_font
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 35

r = 3
for c, h in enumerate(['Feature', 'Descripcion', 'Esfuerzo', 'Notas'], 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 4, BLUE)

p3 = [
    ('Seasons Admin UI', 'Crear/iniciar/completar temporadas desde admin (hoy solo API). Historial de seasons.', '2-3 dias', 'Backend ya implementado, falta UI'),
    ('Double Elimination', 'Winners bracket + losers bracket. Grand final con potential reset.', '4-5 dias', 'Formato standard esports'),
    ('Clanes / Equipos', 'Crear clan (max 5-8), ranking de clanes, tag en alias, torneos 3v3.', '5-7 dias', 'Social bonds: retencion +30%'),
    ('Ligas por Nivel', 'Bronce/Plata/Oro/Diamante. Ascenso/descenso por temporada.', '5-7 dias', 'Requiere Seasons'),
    ('Stats Avanzados', 'Grafico estrellas en tiempo, heatmap actividad, radar chart comparativo.', '3-4 dias', 'SVG puro sin dependencias'),
    ('Hype Mode Torneos', 'Dashboard cambia en torneo vivo. Banner rojo, cheers de espectadores.', '3 dias', 'Espectadores activos'),
    ('Discord Webhook', 'Auto-postear resultados y badges al Discord de la comunidad.', '0.5 dias', 'Solo POST al webhook'),
    ('Admin Analytics', 'DAU/WAU/MAU, partidas/semana, revenue, engagement.', '1-2 dias', 'Data ya existe'),
    ('Meta-game Analytics', 'Win rate por combo, counter matrix, trending combos.', '3-4 dias', 'Registrar combo por partida'),
    ('Ascenso X', 'Rangos X-F-E-D-C-B-A-S-Omega. Torneo final por threshold de estrellas.', '5-7 dias', 'Referencia en Desktop/ascenso X'),
    ('PWA / Offline', 'Service worker, cache offline, "Agregar a inicio".', '2-3 dias', 'Crucial en torneos sin WiFi'),
]

r = 4
for feat in p3:
    for c, val in enumerate(feat, 1):
        ws3.cell(r, c, val)
    style_row(ws3, r, 4, LIGHT_BLUE if r % 2 == 0 else None)
    r += 1

wb.save('Copa_Omega_Plan.xlsx')
print('Excel guardado en Copa_Omega_Plan.xlsx')
